"""列表类 Excel 导出：表头样式、冻结首行、自动列宽。"""

from datetime import datetime
from io import BytesIO
from typing import Any, Sequence
from urllib.parse import quote

from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "微软雅黑"
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="303133")
BODY_FONT = Font(name=FONT_NAME, size=10, color="303133")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="F5F7FA")
THIN_BORDER = Border(
    left=Side(style="thin", color="DCDFE6"),
    right=Side(style="thin", color="DCDFE6"),
    top=Side(style="thin", color="DCDFE6"),
    bottom=Side(style="thin", color="DCDFE6"),
)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
BODY_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def cell_text(value: Any, joiner: str = "、") -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return joiner.join(str(item).strip() for item in value if str(item or "").strip())
    return str(value).strip()


def excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (int, float)):
        return value
    return cell_text(value)


def _cell_display_width(value: Any) -> float:
    if value is None:
        return 0.0
    if hasattr(value, "strftime"):
        return float(len(value.strftime("%Y-%m-%d"))) * 1.25
    width = 0.0
    for ch in str(value):
        # 微软雅黑数字/英文比 Excel 默认列宽单位略宽，纯数字串按 1.0 会裁切
        width += 2.0 if ord(ch) > 127 else 1.25
    return width


def _typical_column_width(widths: list[float]) -> float:
    if not widths:
        return 0.0
    if len(widths) < 50:
        return max(widths)
    widths.sort()
    drop = max(1, len(widths) // 20)
    return widths[-1 - drop]


def auto_fit_worksheet_columns(
    ws,
    *,
    min_width: float = 10.0,
    max_width: float = 60.0,
    padding: float = 2.0,
) -> None:
    for col_idx in range(1, ws.max_column + 1):
        widths = [
            _cell_display_width(ws.cell(row=row_idx, column=col_idx).value)
            for row_idx in range(1, ws.max_row + 1)
        ]
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = min(
            max(_typical_column_width(widths) + padding, min_width),
            max_width,
        )


def build_list_workbook(
    *,
    sheet_title: str,
    filename_prefix: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
) -> tuple[BytesIO, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "导出")[:31]
    ws.append(list(headers))
    ws.row_dimensions[1].height = 22
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    for row in rows:
        ws.append([excel_value(item) for item in row])
        ws.row_dimensions[ws.max_row].height = 18

    if ws.max_row >= 2:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.font = BODY_FONT
                cell.alignment = BODY_ALIGN
                cell.border = THIN_BORDER
                if isinstance(cell.value, str):
                    cell.number_format = "@"

    ws.freeze_panes = "A2"
    auto_fit_worksheet_columns(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return output, filename


def xlsx_response(output: BytesIO, filename: str) -> Response:
    return Response(
        content=output.getvalue(),
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )
