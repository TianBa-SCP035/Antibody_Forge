from collections import defaultdict
from datetime import datetime
from io import BytesIO
from pathlib import Path
import shutil
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from core.config import get_settings
from core.errors import BusinessError

SERUM_CAGE_NO_MOUSE = "SERUM_CAGE_NO_MOUSE"
TERMINAL_PROJECT_STATUSES = frozenset({"结题", "无效价处死"})
STATUS_AUTO_UPDATE_SKIP_STATUSES = TERMINAL_PROJECT_STATUSES | {"加免中"}
from models.immunology import (
    SerumElisaPlate,
    SerumFacsPlate,
    SerumFile,
    SerumImmAntigen,
    SerumImmMouse,
    SerumImmProject,
    SerumImmStep,
    SerumTiterOrder,
    SerumTiterPc,
    SerumTiterTarget,
)
from modules.immunology.titer.service import (
    PENDING_BLOOD_COLLECTION_STATUS,
    _normalize_owner_names,
    create_titer_order_for_blood_collection_if_absent,
)


def _compact_identifier(raw: Any) -> str:
    return "".join(str(raw or "").split())


def _normalize_project_identifiers(data: dict[str, Any]) -> None:
    project_code = _compact_identifier(data.get("project_code"))
    if project_code:
        data["project_code"] = project_code
    experiment_id = _compact_identifier(data.get("experiment_id"))
    if experiment_id:
        data["experiment_id"] = experiment_id


def _steps_query(experiment_id: str):
    return (
        select(SerumImmStep)
        .where(SerumImmStep.experiment_id == experiment_id)
        .order_by(SerumImmStep.group_id.asc(), SerumImmStep.sort_order.asc(), SerumImmStep.step_id.asc())
    )


def _collect_titer_owners_by_experiment(db: Session, experiment_ids: list[str]) -> dict[str, list[str]]:
    ids = [str(exp_id).strip() for exp_id in experiment_ids if str(exp_id or "").strip()]
    if not ids:
        return {}
    owners_map: dict[str, list[str]] = {exp_id: [] for exp_id in ids}
    seen_map: dict[str, set[str]] = {exp_id: set() for exp_id in ids}
    for exp_id, owners_raw in db.execute(
        select(SerumTiterOrder.experiment_id, SerumTiterOrder.titer_owners).where(
            SerumTiterOrder.experiment_id.in_(ids)
        )
    ).all():
        exp = str(exp_id or "").strip()
        if exp not in owners_map:
            continue
        for name in _normalize_owner_names(owners_raw):
            if name not in seen_map[exp]:
                seen_map[exp].add(name)
                owners_map[exp].append(name)
    return owners_map


def apply_project_filters(stmt, data: dict[str, Any]):
    p_code = data.get("project_code")
    p_codes = data.get("project_codes")
    p_name = data.get("project_name")
    owner = data.get("owner")
    status = data.get("project_status")
    target_name = data.get("target_name")
    study_type = data.get("study_type")
    pm = data.get("pm")
    mouse_strain = data.get("mouse_strain")
    mouse_strain_category = data.get("mouse_strain_category")

    stmt = stmt.where(SerumImmProject.project_status != "deleted")
    if p_codes:
        stmt = stmt.where(SerumImmProject.project_code.in_(p_codes))
    elif p_code:
        stmt = stmt.where(SerumImmProject.project_code.like(f"%{p_code}%"))
    if p_name:
        stmt = stmt.where(SerumImmProject.project_name.like(f"%{p_name}%"))
    if owner:
        stmt = stmt.where(SerumImmProject.owner == owner)
    if status:
        if status == "ongoing":
            stmt = stmt.where(
                or_(
                    SerumImmProject.project_status.like("%待%"),
                    SerumImmProject.project_status.like("%已%"),
                    SerumImmProject.project_status == "加免中",
                )
            )
        elif status == "completed":
            stmt = stmt.where(
                SerumImmProject.project_status.in_(("无效价处死", "结题"))
            )
        else:
            stmt = stmt.where(SerumImmProject.project_status == status)
    if target_name:
        stmt = stmt.where(SerumImmProject.target_name == target_name)
    if study_type:
        stmt = stmt.where(SerumImmProject.study_type == study_type)
    if pm:
        stmt = stmt.where(SerumImmProject.pm == pm)
    if mouse_strain:
        stmt = stmt.where(SerumImmProject.mouse_strain == mouse_strain)
    if mouse_strain_category:
        stmt = stmt.where(SerumImmProject.mouse_strain_category == mouse_strain_category)
    return stmt


def get_stats(db: Session) -> dict:
    total = db.scalar(select(func.count(SerumImmProject.id)).where(SerumImmProject.project_status != "deleted")) or 0
    status_counts = db.execute(
        select(SerumImmProject.project_status, func.count(SerumImmProject.id))
        .where(SerumImmProject.project_status != "deleted")
        .group_by(SerumImmProject.project_status)
    ).all()
    status_dict = {status: count for status, count in status_counts}
    ongoing_count = sum(
        count
        for status, count in status_dict.items()
        if status and ("待" in status or "已" in status or status == "加免中")
    )
    completed_count = sum(
        count for status, count in status_dict.items() if status in {"无效价处死", "结题"}
    )
    owner_counts = db.execute(
        select(SerumImmProject.owner, func.count(SerumImmProject.id))
        .where(SerumImmProject.project_status != "deleted")
        .group_by(SerumImmProject.owner)
    ).all()
    return {
        "total": total,
        "status_counts": {**status_dict, "ongoing": ongoing_count, "completed": completed_count},
        "owner_counts": [{"name": owner or "Unknown", "value": count} for owner, count in owner_counts],
    }


def generate_next_id(db: Session, project_code: str) -> str | None:
    project_code = _compact_identifier(project_code)
    if not project_code:
        return None
    projects = db.scalars(
        select(SerumImmProject).where(SerumImmProject.experiment_id.like(f"{project_code}%"))
    ).all()
    existing_suffixes = set()
    for project in projects:
        try:
            suffix = (project.experiment_id or "")[len(project_code) :]
            if suffix.isdigit():
                existing_suffixes.add(int(suffix))
        except Exception:
            continue
    next_suffix = 1
    while next_suffix in existing_suffixes:
        next_suffix += 1
    return f"{project_code}{next_suffix:02d}"


def _project_list_stmt(data: dict[str, Any]):
    stmt = apply_project_filters(select(SerumImmProject), data)
    if data.get("start_date"):
        stmt = stmt.where(SerumImmProject.start_date >= data["start_date"])
    if data.get("end_date"):
        stmt = stmt.where(SerumImmProject.start_date <= data["end_date"])
    return stmt


def _first_cage_by_experiment(db: Session, experiment_ids: list[str]) -> dict[str, str]:
    ids = [str(exp_id).strip() for exp_id in experiment_ids if str(exp_id or "").strip()]
    if not ids:
        return {}
    cages: dict[str, str] = {}
    for exp_id, cage in db.execute(
        select(SerumImmMouse.experiment_id, SerumImmMouse.cage_position)
        .where(
            SerumImmMouse.experiment_id.in_(ids),
            SerumImmMouse.cage_position.is_not(None),
        )
        .order_by(SerumImmMouse.id.asc())
    ).all():
        key = str(exp_id or "").strip()
        if key and key not in cages:
            cages[key] = str(cage or "").strip()
    return cages


def get_list(db: Session, data: dict[str, Any]) -> dict:
    page = int(data.get("page", 1) or 1)
    limit = int(data.get("limit", 20) or 20)
    stmt = _project_list_stmt(data)

    total_stmt = select(func.count()).select_from(stmt.subquery())
    total = db.scalar(total_stmt) or 0
    projects = db.scalars(stmt.order_by(SerumImmProject.id.desc()).offset((page - 1) * limit).limit(limit)).all()
    titer_owners_map = _collect_titer_owners_by_experiment(
        db, [project.experiment_id for project in projects if project.experiment_id]
    )

    items = []
    for project in projects:
        item = project.to_dict()
        item["titer_owners"] = titer_owners_map.get(project.experiment_id or "", [])
        mouse = db.scalar(
            select(SerumImmMouse)
            .where(SerumImmMouse.experiment_id == project.experiment_id, SerumImmMouse.cage_position.is_not(None))
            .limit(1)
        )
        item["cage_position"] = getattr(mouse, "cage_position", "") or ""
        item["cage_position_display"] = item["cage_position"]
        items.append(item)
    return {"items": items, "total": total}


def export_list_workbook(db: Session, data: dict[str, Any]) -> tuple[BytesIO, str]:
    from utils.excel import build_list_workbook

    projects = db.scalars(_project_list_stmt(data or {}).order_by(SerumImmProject.id.desc())).all()
    cages = _first_cage_by_experiment(
        db, [project.experiment_id for project in projects if project.experiment_id]
    )
    headers = [
        "编号",
        "实验ID",
        "项目名称",
        "归类鼠型",
        "笼位",
        "实验备注",
        "课题类型",
        "PM",
        "鼠型",
        "靶点",
        "靶点类型",
        "靶点大小",
        "负责人",
        "开始日期",
        "免疫间隔",
        "检测方法",
        "FACS板数",
        "ELISA板数",
        "状态",
        "制备状态",
        "项目目的",
    ]
    rows = []
    for project in projects:
        rows.append([
            project.project_code,
            project.experiment_id,
            project.project_name,
            project.mouse_strain_category,
            cages.get(str(project.experiment_id or "").strip(), ""),
            project.remark,
            project.study_type,
            project.pm,
            project.mouse_strain,
            project.target_name,
            project.target_type,
            project.target_size,
            project.owner,
            project.start_date,
            project.immunization_interval,
            project.assay_method,
            project.facs_plate_count,
            project.elisa_plate_count,
            project.project_status,
            project.prep_status,
            project.project_purpose,
        ])
    return build_list_workbook(
        sheet_title="免疫实验列表",
        filename_prefix="免疫实验列表",
        headers=headers,
        rows=rows,
    )


def get_detail(db: Session, project_id: int) -> dict | None:
    project = db.get(SerumImmProject, project_id)
    if not project:
        return None
    exp_id = project.experiment_id
    data = project.to_dict()
    data["mouse_groups"] = [item.to_dict() for item in db.scalars(select(SerumImmMouse).where(SerumImmMouse.experiment_id == exp_id)).all()]
    data["antigens"] = [item.to_dict() for item in db.scalars(select(SerumImmAntigen).where(SerumImmAntigen.experiment_id == exp_id)).all()]
    data["steps"] = [item.to_dict() for item in db.scalars(_steps_query(exp_id)).all()]
    data["titer_pcs"] = [item.to_dict() for item in db.scalars(select(SerumTiterPc).where(SerumTiterPc.experiment_id == exp_id)).all()]
    data["titer_targets"] = [item.to_dict() for item in db.scalars(select(SerumTiterTarget).where(SerumTiterTarget.experiment_id == exp_id)).all()]
    data["titer_owners"] = _collect_titer_owners_by_experiment(db, [exp_id]).get(exp_id, [])
    return data


def get_mouse_groups(db: Session, experiment_id: str) -> list[dict]:
    normalized = str(experiment_id or "").strip()
    if not normalized:
        raise ValueError("experiment_id 不能为空")
    rows = db.scalars(
        select(SerumImmMouse)
        .where(SerumImmMouse.experiment_id == normalized)
        .order_by(SerumImmMouse.id.asc())
    ).all()
    return [item.to_dict() for item in rows]


def save_mouse_registry(db: Session, data: dict[str, Any]) -> dict:
    experiment_id = str(data.get("experiment_id") or "").strip()
    if not experiment_id:
        raise ValueError("experiment_id 不能为空")

    row = None
    raw_id = data.get("id")
    if raw_id not in (None, ""):
        row = db.scalar(
            select(SerumImmMouse).where(
                SerumImmMouse.id == int(raw_id),
                SerumImmMouse.experiment_id == experiment_id,
            )
        )
    if row is None:
        group_id = str(data.get("group_id") or "").strip()
        if not group_id:
            raise ValueError("缺少 id 或 group_id")
        row = db.scalar(
            select(SerumImmMouse).where(
                SerumImmMouse.experiment_id == experiment_id,
                SerumImmMouse.group_id == group_id,
            )
        )
    if not row:
        raise ValueError("小鼠分组不存在")

    if "mouse_registry" in data:
        row.mouse_registry = data.get("mouse_registry")
    if "mouse_no_list" in data:
        row.mouse_no_list = str(data.get("mouse_no_list") or "")
    db.commit()
    db.refresh(row)
    return row.to_dict()


def _update_fields(obj, item: dict[str, Any], skip: set[str]) -> None:
    for key, value in item.items():
        if key not in skip and hasattr(obj, key):
            setattr(obj, key, value)


def incremental_update(db: Session, model_class, items: list[dict], experiment_id: str, id_field: str = "id") -> list:
    submitted_ids = set()
    new_items = []
    for item in items:
        item_id = item.get(id_field)
        if item_id not in (None, ""):
            obj = db.get(model_class, int(item_id))
            if obj:
                _update_fields(obj, item, {id_field})
                obj.experiment_id = experiment_id
                submitted_ids.add(int(item_id))
            else:
                new_items.append(item)
        else:
            new_items.append(item)

    existing = db.scalars(select(model_class).where(model_class.experiment_id == experiment_id)).all()
    for obj in existing:
        if submitted_ids and getattr(obj, id_field) not in submitted_ids:
            db.delete(obj)
        elif not submitted_ids:
            db.delete(obj)

    inserted = []
    for item in new_items:
        item_data = {key: value for key, value in item.items() if key != id_field}
        item_data["experiment_id"] = experiment_id
        obj = model_class(**item_data)
        db.add(obj)
        inserted.append(obj)
    return inserted


def bulk_insert(db: Session, model_class, items: list[dict], experiment_id: str, id_field: str = "id") -> list:
    inserted = []
    for item in items:
        item_data = {key: value for key, value in item.items() if key != id_field}
        item_data["experiment_id"] = experiment_id
        obj = model_class(**item_data)
        db.add(obj)
        inserted.append(obj)
    return inserted


PROJECT_FIELDS = [
    "experiment_id",
    "project_code",
    "project_name",
    "project_purpose",
    "start_date",
    "immunization_interval",
    "target_name",
    "target_type",
    "target_size",
    "owner",
    "pm",
    "study_type",
    "assay_method",
    "facs_plate_count",
    "elisa_plate_count",
    "project_status",
    "remark",
]


def _titer_upload_root() -> Path:
    return Path(get_settings().repository_root) / "uploads" / "titer_files"


def _next_available_path(path: Path) -> Path:
    if not path.exists():
        return path
    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    candidate = path.with_name(f"{path.stem}_{stamp}{path.suffix}")
    index = 1
    while candidate.exists():
        candidate = path.with_name(f"{path.stem}_{stamp}_{index}{path.suffix}")
        index += 1
    return candidate


def _move_titer_experiment_dir(old_eid: str, new_eid: str) -> dict[str, str]:
    upload_root = _titer_upload_root()
    old_dir = upload_root / old_eid
    new_dir = upload_root / new_eid
    path_map: dict[str, str] = {}
    if not old_dir.exists():
        return path_map
    if not new_dir.exists():
        new_dir.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(old_dir), str(new_dir))
        return path_map

    for source in old_dir.rglob("*"):
        if not source.is_file():
            continue
        relative = source.relative_to(old_dir).as_posix()
        target = _next_available_path(new_dir / relative)
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(source), str(target))
        path_map[relative] = f"/titer_files/{new_eid}/{target.relative_to(new_dir).as_posix()}"
    return path_map


def _rewrite_titer_file_path(file_path: str | None, old_eid: str, new_eid: str, path_map: dict[str, str]) -> str | None:
    if not file_path:
        return file_path
    old_prefix = f"/titer_files/{old_eid}/"
    if not file_path.startswith(old_prefix):
        return file_path
    relative = file_path[len(old_prefix) :]
    return path_map.get(relative, f"/titer_files/{new_eid}/{relative}")


def _rename_experiment_related_records(db: Session, old_eid: str | None, new_eid: str | None) -> None:
    if not old_eid or not new_eid or old_eid == new_eid:
        return
    path_map = _move_titer_experiment_dir(old_eid, new_eid)
    for record in db.scalars(select(SerumFile).where(SerumFile.experiment_id == old_eid)).all():
        record.file_path = _rewrite_titer_file_path(record.file_path, old_eid, new_eid, path_map)
        record.experiment_id = new_eid
    for model in [
        SerumImmMouse,
        SerumImmAntigen,
        SerumImmStep,
        SerumTiterTarget,
        SerumTiterPc,
        SerumFacsPlate,
        SerumElisaPlate,
        SerumTiterOrder,
    ]:
        db.query(model).filter(model.experiment_id == old_eid).update({"experiment_id": new_eid}, synchronize_session=False)


def save_serum(db: Session, data: dict[str, Any]) -> dict:
    _normalize_project_identifiers(data)
    project_id = data.get("id")
    new_eid = data.get("experiment_id")
    new_mice = new_antigens = new_steps = new_targets = new_pcs = []

    if project_id:
        project = db.get(SerumImmProject, int(project_id))
        if not project:
            raise ValueError("项目不存在")
        old_eid = project.experiment_id
        if old_eid != new_eid:
            if not new_eid:
                raise ValueError("实验 ID 不能为空")
            existing_project = db.scalar(
                select(SerumImmProject).where(
                    SerumImmProject.experiment_id == new_eid,
                    SerumImmProject.id != project.id,
                )
            )
            if existing_project:
                raise ValueError("实验 ID 已存在")
        for field in PROJECT_FIELDS:
            setattr(project, field, data.get(field))
        if old_eid != new_eid:
            _rename_experiment_related_records(db, old_eid, new_eid)
        new_mice = incremental_update(db, SerumImmMouse, data.get("mouse_groups", []), new_eid)
        new_antigens = incremental_update(db, SerumImmAntigen, data.get("antigens", []), new_eid)
        new_steps = incremental_update(db, SerumImmStep, data.get("steps", []), new_eid, id_field="step_id")
        new_targets = incremental_update(db, SerumTiterTarget, data.get("titer_targets", []), new_eid)
        new_pcs = incremental_update(db, SerumTiterPc, data.get("titer_pcs", []), new_eid)
    else:
        if not new_eid:
            new_eid = generate_next_id(db, data.get("project_code"))
            if not new_eid:
                raise ValueError("项目编号不能为空，无法生成实验 ID")
        project = SerumImmProject(**{field: data.get(field) for field in PROJECT_FIELDS})
        project.experiment_id = new_eid
        db.add(project)
        db.flush()
        new_mice = bulk_insert(db, SerumImmMouse, data.get("mouse_groups", []), new_eid)
        new_antigens = bulk_insert(db, SerumImmAntigen, data.get("antigens", []), new_eid)
        new_steps = bulk_insert(db, SerumImmStep, data.get("steps", []), new_eid, id_field="step_id")
        new_targets = bulk_insert(db, SerumTiterTarget, data.get("titer_targets", []), new_eid)
        new_pcs = bulk_insert(db, SerumTiterPc, data.get("titer_pcs", []), new_eid)

    mouse_groups = data.get("mouse_groups", [])
    project.mouse_strain = "+".join(sorted({m.get("mouse_strain", "").strip() for m in mouse_groups if m.get("mouse_strain")}))
    project.mouse_strain_category = "+".join(sorted({m.get("mouse_strain_category", "").strip() for m in mouse_groups if m.get("mouse_strain_category")}))
    db.commit()

    response = {
        "id": project.id,
        "experiment_id": project.experiment_id,
        "project_code": project.project_code,
    }
    if new_mice:
        response["new_mouse_records"] = [item.to_dict() for item in new_mice]
    if new_antigens:
        response["new_antigen_records"] = [item.to_dict() for item in new_antigens]
    if new_steps:
        response["new_step_records"] = [item.to_dict() for item in new_steps]
    if new_targets:
        response["new_target_records"] = [item.to_dict() for item in new_targets]
    if new_pcs:
        response["new_pc_records"] = [item.to_dict() for item in new_pcs]
    return response


def delete_serum(db: Session, project_id: int) -> None:
    project = db.get(SerumImmProject, project_id)
    if not project:
        raise ValueError("项目不存在")
    exp_id = project.experiment_id
    for model in [
        SerumFacsPlate,
        SerumElisaPlate,
        SerumFile,
        SerumImmMouse,
        SerumImmAntigen,
        SerumImmStep,
        SerumTiterTarget,
        SerumTiterPc,
        SerumTiterOrder,
    ]:
        db.query(model).filter(model.experiment_id == exp_id).delete(synchronize_session=False)
    db.delete(project)
    db.commit()


def update_status(db: Session, project_id: int, project_status: str) -> None:
    project = db.get(SerumImmProject, project_id)
    if not project:
        raise ValueError("项目不存在")
    project.project_status = project_status
    db.commit()


def update_cage_position(db: Session, project_id: int, cage_position: str | None) -> None:
    project = db.get(SerumImmProject, project_id)
    if not project:
        raise ValueError("项目不存在")
    count = db.scalar(select(func.count(SerumImmMouse.id)).where(SerumImmMouse.experiment_id == project.experiment_id)) or 0
    if count == 0:
        raise BusinessError("鼠鼠不存在", error_code=SERUM_CAGE_NO_MOUSE)
    db.query(SerumImmMouse).filter(SerumImmMouse.experiment_id == project.experiment_id).update(
        {"cage_position": cage_position},
        synchronize_session=False,
    )
    db.commit()


def update_prep_status(db: Session, experiment_id: str, prep_status: str | None) -> None:
    project = db.scalar(select(SerumImmProject).where(SerumImmProject.experiment_id == experiment_id))
    if not project:
        raise ValueError("项目不存在")
    project.prep_status = prep_status
    db.commit()


def _next_scheduled_status(rec: dict[str, Any], today: str) -> str | None:
    if today < rec["min_d"]:
        return f"待{rec['min_stage']}"
    if rec.get("next_stage"):
        return f"待{rec['next_stage']}"
    return None


def _build_immunization_schedule(db: Session, experiment_ids: list[str], today: str) -> dict[str, dict[str, Any]]:
    if not experiment_ids:
        return {}
    steps = db.execute(
        select(SerumImmStep.experiment_id, SerumImmStep.date_actual, SerumImmStep.stage_name).where(
            SerumImmStep.experiment_id.in_(experiment_ids),
            SerumImmStep.date_actual.is_not(None),
            SerumImmStep.date_actual != "",
            SerumImmStep.stage_name.is_not(None),
            SerumImmStep.stage_name != "",
        )
    )
    info: dict[str, dict[str, Any]] = {}
    for exp, date_value, stage in steps:
        date_value = (date_value or "").strip()
        stage = (stage or "").strip()
        rec = info.setdefault(
            exp,
            {"min_d": date_value, "min_stage": stage, "max_d": date_value, "next_d": None, "next_stage": None},
        )
        if date_value < rec["min_d"]:
            rec["min_d"], rec["min_stage"] = date_value, stage
        if date_value > rec["max_d"]:
            rec["max_d"] = date_value
        if date_value > today and (rec["next_d"] is None or date_value < rec["next_d"]):
            rec["next_d"], rec["next_stage"] = date_value, stage
    return info


def auto_update_status(db: Session, filters: dict[str, Any] | None = None) -> dict:
    filters = filters or {}
    proj_stmt = apply_project_filters(select(SerumImmProject), filters)
    if filters.get("start_date"):
        proj_stmt = proj_stmt.where(SerumImmProject.start_date >= filters["start_date"])
    if filters.get("end_date"):
        proj_stmt = proj_stmt.where(SerumImmProject.start_date <= filters["end_date"])
    projects = db.scalars(proj_stmt).all()
    if not projects:
        return {"message": "未找到符合条件的项目", "updated_count": 0, "titer_order_created_count": 0}

    today = datetime.now().strftime("%Y-%m-%d")
    schedule_exp_ids = [
        project.experiment_id
        for project in projects
        if project.experiment_id and (project.project_status or "") not in TERMINAL_PROJECT_STATUSES
    ]
    schedule = _build_immunization_schedule(db, schedule_exp_ids, today)

    updated_count = 0
    titer_order_created_count = 0
    dry_run = bool(filters.get("dry_run"))
    for project in projects:
        if not project.experiment_id:
            continue
        status = project.project_status or ""
        rec = schedule.get(project.experiment_id)
        if not rec or today >= rec["max_d"]:
            continue

        next_status = _next_scheduled_status(rec, today)

        if status not in TERMINAL_PROJECT_STATUSES:
            if not dry_run and next_status == PENDING_BLOOD_COLLECTION_STATUS and rec.get("next_d"):
                if create_titer_order_for_blood_collection_if_absent(db, project, rec["next_d"]):
                    titer_order_created_count += 1

        if status in STATUS_AUTO_UPDATE_SKIP_STATUSES:
            continue
        if next_status and project.project_status != next_status:
            if not dry_run:
                project.project_status = next_status
            updated_count += 1
    if dry_run:
        db.rollback()
    else:
        db.commit()
    return {
        "message": "状态更新成功",
        "updated_count": updated_count,
        "titer_order_created_count": titer_order_created_count,
    }


def get_filter_options(db: Session) -> dict:
    def distinct_values(column):
        rows = db.execute(select(column).distinct().where(column.is_not(None), column != "").order_by(column)).all()
        return [row[0] for row in rows if row[0]]

    return {
        "targets": distinct_values(SerumImmProject.target_name),
        "owners": distinct_values(SerumImmProject.owner),
        "study_types": distinct_values(SerumImmProject.study_type),
        "pms": distinct_values(SerumImmProject.pm),
        "mouse_strains": distinct_values(SerumImmProject.mouse_strain),
        "mouse_strain_categories": distinct_values(SerumImmProject.mouse_strain_category),
        "statuses": distinct_values(SerumImmProject.project_status),
    }


def _cell_display_width(value: Any) -> float:
    if value is None:
        return 0.0
    if hasattr(value, "strftime"):
        return float(len(value.strftime("%Y-%m-%d")))
    width = 0.0
    for ch in str(value):
        width += 2.0 if ord(ch) > 127 else 1.0
    return width


def _auto_fit_worksheet_columns(
    ws,
    *,
    min_width: float = 8.0,
    max_width: float = 50.0,
    padding: float = 2.0,
) -> None:
    for col_idx in range(1, ws.max_column + 1):
        content_width = max(
            (_cell_display_width(ws.cell(row=row_idx, column=col_idx).value) for row_idx in range(1, ws.max_row + 1)),
            default=0.0,
        )
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = min(max(content_width + padding, min_width), max_width)


def export_mouse_workbook(db: Session, data: dict[str, Any]) -> tuple[BytesIO, str]:
    stmt = apply_project_filters(select(SerumImmProject), data)
    projects = db.scalars(stmt.order_by(SerumImmProject.id.desc())).all()
    filename = f"小鼠免疫导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "血清实验数据导出"
    headers = ["项目编号", "实际日期", "靶点", "鼠型", "组别", "只数", "笼位", "抗原种属", "抗原类型", "抗原名称", "原液浓度", "剂量", "给药途径", "免疫阶段", "免疫备注", "备注", "免疫负责人"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="微软雅黑", size=11, bold=True)
    ws.freeze_panes = "A2"

    exp_ids = [project.experiment_id for project in projects if project.experiment_id]
    if exp_ids:
        all_mice = db.scalars(select(SerumImmMouse).where(SerumImmMouse.experiment_id.in_(exp_ids))).all()
        all_antigens = db.scalars(select(SerumImmAntigen).where(SerumImmAntigen.experiment_id.in_(exp_ids))).all()
        steps_stmt = select(SerumImmStep).where(SerumImmStep.experiment_id.in_(exp_ids))
        if data.get("start_date"):
            steps_stmt = steps_stmt.where(SerumImmStep.date_actual >= data["start_date"])
        if data.get("end_date"):
            steps_stmt = steps_stmt.where(SerumImmStep.date_actual <= data["end_date"])
        all_steps = db.scalars(steps_stmt.order_by(SerumImmStep.date_actual.asc())).all()

        mice_by_exp = defaultdict(list)
        for mouse in all_mice:
            mice_by_exp[mouse.experiment_id].append(mouse)
        antigens_by_exp = defaultdict(dict)
        for antigen in all_antigens:
            antigens_by_exp[antigen.experiment_id][antigen.antigen_id] = antigen
        steps_by_exp = defaultdict(lambda: defaultdict(list))
        for step in all_steps:
            steps_by_exp[step.experiment_id][step.group_id].append(step)

        for project in projects:
            for mouse in mice_by_exp.get(project.experiment_id, []):
                for step in steps_by_exp.get(project.experiment_id, {}).get(mouse.group_id, []):
                    antigen_info = _resolve_antigen_info(antigens_by_exp.get(project.experiment_id, {}), step.antigen_id)
                    ws.append([
                        project.project_code,
                        step.date_actual,
                        project.target_name,
                        mouse.mouse_strain,
                        mouse.group_id,
                        mouse.mouse_count,
                        mouse.cage_position,
                        antigen_info.get("species", ""),
                        antigen_info.get("antigen_type", ""),
                        antigen_info.get("antigen_name", ""),
                        antigen_info.get("stock_conc", ""),
                        step.antigen_dose,
                        step.route,
                        step.stage_name,
                        step.remark,
                        project.remark,
                        project.owner or "",
                    ])

    _auto_fit_worksheet_columns(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, filename


def _resolve_antigen_info(antigen_dict: dict, antigen_id: str | None) -> dict:
    if not antigen_id:
        return {}
    if "," not in str(antigen_id):
        antigen = antigen_dict.get(antigen_id)
        return antigen.to_dict() if antigen else {}
    antigen_ids = [item.strip() for item in str(antigen_id).split(",") if item.strip()]
    antigens = [antigen_dict.get(item) for item in antigen_ids if antigen_dict.get(item)]
    return {
        "species": " + ".join([item.species or "" for item in antigens]),
        "antigen_type": " + ".join([item.antigen_type or "" for item in antigens]),
        "antigen_name": " + ".join([item.antigen_name or "" for item in antigens]),
        "stock_conc": " + ".join([item.stock_conc or "" for item in antigens]),
    }
