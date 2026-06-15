"""免疫方案 Excel 导出：单项目工作簿构建 + 单/批量统一响应。"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any
import math
import subprocess
import sys
import tempfile
import zipfile

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from models.immunology import SerumImmAntigen, SerumImmMouse, SerumImmProject, SerumImmStep

from core.config import get_settings

# A1 留边距（仅屏幕阅读）；正文 B2:P*。B=侧栏，C~P=数据区，宽列 D:H。
MARGIN_ROW = 1
MARGIN_COL = 1
CONTENT_START_ROW = 2
SECTION_COL = 2
DATA_START_COL = 3
LAST_COL = 16

FONT_NAME = "微软雅黑"
FONT_BODY = Font(name=FONT_NAME, size=10)
FONT_BOLD = Font(name=FONT_NAME, size=10, bold=True)
FONT_TITLE = Font(name=FONT_NAME, size=15, bold=True)
FONT_SIDEBAR = Font(name=FONT_NAME, size=11, bold=True)

CELL_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
LABEL_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="E8E8E8")
GAP_FILL = PatternFill(fill_type="solid", fgColor="FAFAFA")

WIDE_START = 4
WIDE_END = 8
STEP_ANTIGEN_START = 6
STEP_ANTIGEN_END = 9

DEFAULT_ROW_HEIGHT = 20.0
MIN_COL_WIDTH = 6.0
MAX_COL_WIDTH = 42.0

# 打印页边距（英寸）：介于 Excel「窄」与「常规」之间（常规左右 0.75 / 上下 1.0）
PRINT_MARGIN_LEFT = 0.5
PRINT_MARGIN_RIGHT = 0.5
PRINT_MARGIN_TOP = 0.8
PRINT_MARGIN_BOTTOM = 0.8
PRINT_MARGIN_HEADER = 0.3
PRINT_MARGIN_FOOTER = 0.3
# Excel 纸张代码：9=A4，13=ISO B5（176×250mm）；写入文件后打印预览默认跟此走
PRINT_PAPER_SIZE = 13

CATALOG_COLS = (11, 12)
MOUSE_STRAIN_COLS = (4, 5)
MOUSE_NO_COLS = range(12, 16)
PLAN_DATE_COL = 4  # D 列：步骤表「计划日期」

TableCol = tuple[int, int, str]

ANTIGEN_COLUMNS: list[TableCol] = [
    (DATA_START_COL, DATA_START_COL, "抗原ID"),
    (WIDE_START, WIDE_END, "抗原名称"),
    (9, 9, "抗原种属"),
    (10, 10, "抗原类型"),
    (11, 11, "货号"),
    (12, 12, "批号"),
    (13, 13, "原液浓度"),
    (14, 14, "供应商"),
    (15, 15, "佐剂类型"),
    (16, 16, "佐剂来源"),
]

MOUSE_COLUMNS: list[TableCol] = [
    (DATA_START_COL, DATA_START_COL, "组别"),
    (4, 5, "小鼠名称/品系"),
    (6, 6, "归类鼠型"),
    (7, 7, "免疫数量"),
    (8, 8, "周龄"),
    (9, 9, "性别"),
    (10, 10, "笼位"),
    (11, 11, "供应商"),
    (12, 15, "鼠号列表"),
    (16, 16, "备注"),
]

STEP_COLUMNS: list[TableCol] = [
    (DATA_START_COL, DATA_START_COL, "阶段"),
    (4, 4, "计划日期"),
    (5, 5, "相对天数"),
    (STEP_ANTIGEN_START, STEP_ANTIGEN_END, "抗原"),
    (10, 10, "剂量"),
    (11, 11, "佐剂"),
    (12, 12, "CPG剂量"),
    (13, 13, "注射体积"),
    (14, 14, "途径"),
    (15, 15, "注射部位"),
    (16, 16, "备注"),
]


def load_scheme_data(db: Session, project_id: int) -> dict[str, Any] | None:
    project = db.get(SerumImmProject, project_id)
    if not project:
        return None
    exp_id = project.experiment_id
    return {
        **project.to_dict(),
        "mouse_groups": [
            item.to_dict()
            for item in db.scalars(select(SerumImmMouse).where(SerumImmMouse.experiment_id == exp_id)).all()
        ],
        "antigens": [
            item.to_dict()
            for item in db.scalars(select(SerumImmAntigen).where(SerumImmAntigen.experiment_id == exp_id)).all()
        ],
        "steps": [
            item.to_dict()
            for item in db.scalars(select(SerumImmStep).where(SerumImmStep.experiment_id == exp_id)).all()
        ],
    }


def export_scheme_response(db: Session, project_ids: list[int]) -> tuple[BytesIO, str, str]:
    unique_ids = list(dict.fromkeys(int(pid) for pid in project_ids))
    schemes = [data for pid in unique_ids if (data := load_scheme_data(db, pid))]
    if not schemes:
        raise ValueError("未找到可导出的项目")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if len(schemes) == 1:
        project = schemes[0]
        return _build_project_workbook(project), _scheme_filename(project, "xlsx", timestamp), "xlsx"

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        used_names: set[str] = set()
        for project in schemes:
            code = _safe_filename_part(project.get("project_code") or project.get("experiment_id") or project.get("id"))
            filename = f"免疫方案_{code}.xlsx"
            suffix = 2
            while filename in used_names:
                filename = f"免疫方案_{code}_{suffix}.xlsx"
                suffix += 1
            used_names.add(filename)
            archive.writestr(filename, _build_project_workbook(project).getvalue())

    zip_buffer.seek(0)
    return zip_buffer, f"免疫方案批量导出_{timestamp}.zip", "zip"


def _safe_filename_part(value: Any) -> str:
    text = str(value or "unknown").strip()
    for ch in '\\/:*?"<>|':
        text = text.replace(ch, "_")
    return text or "unknown"


def _scheme_filename(project: dict[str, Any], ext: str, timestamp: str) -> str:
    code = _safe_filename_part(project.get("project_code") or project.get("experiment_id") or project.get("id"))
    return f"免疫方案_{code}_{timestamp}.{ext}"


def _build_project_workbook(project: dict[str, Any]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "免疫方案"
    body_end_row = _SchemeSheetWriter(ws).write_project(project)
    _apply_sheet_borders(ws, CONTENT_START_ROW, body_end_row)
    _autofit_columns(ws)
    _autofit_rows(ws)
    _configure_print(ws, body_end_row)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _merge_bounds(ws: Worksheet, row: int, col: int) -> tuple[int, int, int, int]:
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return merged.min_row, merged.min_col, merged.max_row, merged.max_col
    return row, col, row, col


def _iter_logical_units(ws: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int):
    seen: set[tuple[int, int]] = set()
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            if (row, col) in seen:
                continue
            top, left, bottom, right = _merge_bounds(ws, row, col)
            for r in range(top, bottom + 1):
                for c in range(left, right + 1):
                    if (r, c) != (top, left):
                        seen.add((r, c))
            yield top, left


def _apply_sheet_borders(ws: Worksheet, min_row: int, max_row: int) -> None:
    """拆合 → 正文区逐格 thin → 合并 → 逻辑单元 master 再补一遍。"""
    merges = [
        (m.min_row, m.min_col, m.max_row, m.max_col)
        for m in list(ws.merged_cells.ranges)
        if m.min_row != m.max_row or m.min_col != m.max_col
    ]
    for top, left, bottom, right in merges:
        ws.unmerge_cells(start_row=top, start_column=left, end_row=bottom, end_column=right)

    for row in range(min_row, max_row + 1):
        for col in range(SECTION_COL, LAST_COL + 1):
            ws.cell(row, col).border = CELL_BORDER

    for top, left, bottom, right in merges:
        ws.merge_cells(start_row=top, start_column=left, end_row=bottom, end_column=right)

    for top, left in _iter_logical_units(ws, min_row, max_row, SECTION_COL, LAST_COL):
        ws.cell(top, left).border = CELL_BORDER


def _configure_print(ws: Worksheet, body_end_row: int) -> None:
    ws.print_area = f"{get_column_letter(SECTION_COL)}{CONTENT_START_ROW}:{get_column_letter(LAST_COL)}{body_end_row}"
    ws.page_margins = PageMargins(
        left=PRINT_MARGIN_LEFT,
        right=PRINT_MARGIN_RIGHT,
        top=PRINT_MARGIN_TOP,
        bottom=PRINT_MARGIN_BOTTOM,
        header=PRINT_MARGIN_HEADER,
        footer=PRINT_MARGIN_FOOTER,
    )
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = PRINT_PAPER_SIZE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True


def _text(value: Any, *, empty: str = "-") -> str:
    if value is None:
        return empty
    text = str(value).strip()
    return text if text else empty


def _cell_display_width(value: Any) -> float:
    if value is None:
        return 0.0
    return sum(2.0 if ord(ch) > 127 else 1.0 for ch in str(value))


def _autofit_columns(ws: Worksheet) -> None:
    ws.row_dimensions[MARGIN_ROW].height = 8.0
    ws.column_dimensions[get_column_letter(MARGIN_COL)].width = 2.0
    ws.column_dimensions[get_column_letter(SECTION_COL)].width = 5.0

    for col in range(DATA_START_COL, LAST_COL + 1):
        letter = get_column_letter(col)
        max_width = 0.0
        for row in range(CONTENT_START_ROW, ws.max_row + 1):
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                max_width = max(max_width, _cell_display_width(cell.value))

        if WIDE_START <= col <= WIDE_END:
            cap, floor = MAX_COL_WIDTH, 10.0
        elif col == DATA_START_COL:
            cap, floor = 12.0, MIN_COL_WIDTH
        else:
            cap, floor = 18.0, MIN_COL_WIDTH
        ws.column_dimensions[letter].width = min(max(max_width + 2.0, floor), cap)

    for col in CATALOG_COLS:
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 0, 11.5)
    for col in MOUSE_STRAIN_COLS:
        letter = get_column_letter(col)
        w = ws.column_dimensions[letter].width or MIN_COL_WIDTH
        ws.column_dimensions[letter].width = min(max(w, 7.0), 11.0)
    for col in MOUSE_NO_COLS:
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 0, 9.0)

    for wide_start, wide_end in ((WIDE_START, WIDE_END), (STEP_ANTIGEN_START, STEP_ANTIGEN_END)):
        wide_max = 0.0
        for row in range(CONTENT_START_ROW, ws.max_row + 1):
            for col in range(wide_start, wide_end + 1):
                cell = ws.cell(row, col)
                if isinstance(cell, MergedCell):
                    continue
                _, left, _, right = _merge_bounds(ws, row, col)
                if left == wide_start and right == wide_end and col == left:
                    wide_max = max(wide_max, _cell_display_width(cell.value))
        if wide_max > 0:
            share = min(max(wide_max / (wide_end - wide_start + 1) + 1.5, 10.0), 16.0)
            for col in range(wide_start, wide_end + 1):
                ws.column_dimensions[get_column_letter(col)].width = share

    # D 列兼抗原宽列与步骤「计划日期」，宽列均分后再保底，避免表头「计划日期」换行
    letter = get_column_letter(PLAN_DATE_COL)
    ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 0, 12.0)


def _autofit_rows(ws: Worksheet) -> None:
    for row in range(CONTENT_START_ROW, ws.max_row + 1):
        line_count = 1
        for col in range(SECTION_COL, LAST_COL + 1):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell) or not cell.value or not str(cell.value).strip():
                continue
            if not cell.alignment or not cell.alignment.wrap_text:
                continue
            _, left, _, right = _merge_bounds(ws, row, col)
            if col != left:
                continue
            span_width = sum(ws.column_dimensions[get_column_letter(c)].width or MIN_COL_WIDTH for c in range(left, right + 1))
            if span_width <= 0:
                continue
            text_width = _cell_display_width(cell.value)
            lines = max(1, math.ceil(text_width / max(int(span_width / 2.2), 4)))
            line_count = max(line_count, lines)
        target = max(DEFAULT_ROW_HEIGHT, min(line_count * 15.0, 120.0))
        current = ws.row_dimensions[row].height
        if current is None or target > current:
            ws.row_dimensions[row].height = target


def _resolve_antigen_display(antigen_map: dict[str, dict], antigen_id: str | None) -> str:
    if not antigen_id:
        return "-"
    raw = str(antigen_id).strip()
    if raw.upper() == "N/A":
        return "N/A"
    if "," not in raw:
        antigen = antigen_map.get(raw)
        return _text(antigen.get("antigen_name") if antigen else raw, empty=raw)
    names = []
    for part in raw.split(","):
        key = part.strip()
        if not key:
            continue
        antigen = antigen_map.get(key)
        names.append(_text(antigen.get("antigen_name") if antigen else key, empty=key))
    return " + ".join(names)


def _sort_steps(steps: list[dict[str, Any]]) -> list[dict[str, Any]]:
    def sort_key(step: dict[str, Any]) -> tuple[int, str]:
        day = step.get("day_relative")
        try:
            day_num = int(day)
        except (TypeError, ValueError):
            day_num = 10**9
        return day_num, _text(step.get("date_actual"), empty="")

    return sorted(steps, key=sort_key)


class _SchemeSheetWriter:
    def __init__(self, ws: Worksheet) -> None:
        self.ws = ws
        self.row = CONTENT_START_ROW

    def write_project(self, project: dict[str, Any]) -> int:
        code = _text(project.get("project_code"), empty="")
        title = f"小鼠免疫方案 — {code}" if code else "小鼠免疫方案"
        self._write_title(title)
        self._write_basic_info(project)
        self._write_gap_row()
        self._write_section_table("抗原信息", ANTIGEN_COLUMNS, self._antigen_rows(project.get("antigens") or []))
        self._write_gap_row()
        self._write_section_table("小鼠分组", MOUSE_COLUMNS, self._mouse_rows(project.get("mouse_groups") or []))
        self._write_gap_row()
        self._write_steps_section(project)
        self._write_gap_row()
        self._write_footer()
        return self.row - 1

    def _antigen_rows(self, antigens: list[dict[str, Any]]) -> list[list[str]]:
        return [
            [
                _text(a.get("antigen_id")),
                _text(a.get("antigen_name")),
                _text(a.get("species")),
                _text(a.get("antigen_type")),
                _text(a.get("catalog_no")),
                _text(a.get("lot_no")),
                _text(a.get("stock_conc")),
                _text(a.get("vendor")),
                _text(a.get("adjuvant_type")),
                _text(a.get("adjuvant_source")),
            ]
            for a in antigens
        ]

    def _mouse_rows(self, mouse_groups: list[dict[str, Any]]) -> list[list[str]]:
        return [
            [
                _text(g.get("group_id")),
                _text(g.get("mouse_strain")),
                _text(g.get("mouse_strain_category")),
                _text(g.get("mouse_count")),
                _text(g.get("age_weeks")),
                _text(g.get("sex")),
                _text(g.get("cage_position")),
                _text(g.get("vendor")),
                _text(g.get("mouse_no_list")),
                _text(g.get("remark")),
            ]
            for g in mouse_groups
        ]

    def _write_title(self, title: str) -> None:
        self._merge_write(
            self.row, SECTION_COL, LAST_COL, title,
            font=FONT_TITLE,
            alignment=Alignment(horizontal="center", vertical="center"),
            height=32,
        )
        self.row += 1

    def _write_basic_info(self, project: dict[str, Any]) -> None:
        interval = _text(project.get("immunization_interval"), empty="")
        interval = f"{interval} 天" if interval and interval != "-" else "-"
        rows: list[tuple[str, str, str, str, str, str]] = [
            ("项目编号", _text(project.get("project_code")), "实验ID", _text(project.get("experiment_id")), "负责人", _text(project.get("owner"))),
            ("靶点名称", _text(project.get("target_name")), "靶点类型", _text(project.get("target_type")), "靶点大小", _text(project.get("target_size"))),
            ("项目名称", _text(project.get("project_name")), "课题类型", _text(project.get("study_type")), "PM", _text(project.get("pm"))),
            ("开始日期", _text(project.get("start_date")), "检测方法", _text(project.get("assay_method")), "项目状态", _text(project.get("project_status"))),
            ("免疫间隔", interval, "实验备注", _text(project.get("remark")), "", ""),
        ]
        for labels in rows:
            self._write_kv_row(*labels)
        self._write_single_kv_row("项目目的", _text(project.get("project_purpose")), wrap=True)

    def _write_kv_row(self, label1: str, value1: str, label2: str, value2: str, label3: str, value3: str) -> None:
        col = SECTION_COL
        for label, value in ((label1, value1), (label2, value2), (label3, value3)):
            if not label:
                if col <= LAST_COL:
                    self._merge_write(self.row, col, LAST_COL, "", fill=LABEL_FILL)
                break
            self._merge_write(self.row, col, col + 1, label, font=FONT_BOLD, fill=LABEL_FILL, alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))
            self._merge_write(self.row, col + 2, col + 4, value, alignment=Alignment(horizontal="left", vertical="center", wrap_text=True))
            col += 5
        self.row += 1

    def _write_single_kv_row(self, label: str, value: str, *, wrap: bool = False) -> None:
        self._merge_write(self.row, SECTION_COL, SECTION_COL + 1, label, font=FONT_BOLD, fill=LABEL_FILL, alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))
        self._merge_write(self.row, SECTION_COL + 2, LAST_COL, value, alignment=Alignment(horizontal="left", vertical="top", wrap_text=wrap))
        self.row += 1

    def _write_gap_row(self) -> None:
        self._merge_write(self.row, SECTION_COL, SECTION_COL, "", fill=GAP_FILL, height=8)
        self._merge_write(self.row, DATA_START_COL, LAST_COL, "", fill=GAP_FILL)
        self.row += 1

    def _write_steps_section(self, project: dict[str, Any]) -> None:
        steps_by_group: dict[str, list[dict[str, Any]]] = {}
        for step in project.get("steps") or []:
            steps_by_group.setdefault(_text(step.get("group_id"), empty="UNKNOWN"), []).append(step)

        antigen_map = {
            _text(item.get("antigen_id"), empty=""): item
            for item in (project.get("antigens") or [])
            if item.get("antigen_id") is not None
        }
        mouse_groups = project.get("mouse_groups") or []
        group_order = [_text(g.get("group_id"), empty="") for g in mouse_groups if _text(g.get("group_id"), empty="")]
        for gid in steps_by_group:
            if gid not in group_order:
                group_order.append(gid)

        section_start = self.row
        header_written = False
        if not group_order:
            self._write_table_header_row(STEP_COLUMNS)
            self._write_table_data_row(STEP_COLUMNS, ["-"] * len(STEP_COLUMNS))
        else:
            group_lookup = {_text(g.get("group_id"), empty=""): g for g in mouse_groups}
            for gid in group_order:
                group = group_lookup.get(gid, {})
                strain = _text(group.get("mouse_strain"), empty="")
                count = _text(group.get("mouse_count"), empty="")
                subtitle = f"分组 {gid}"
                if strain:
                    subtitle += f"  ·  {strain}"
                if count and count != "-":
                    subtitle += f"  ·  {count}只"
                self._merge_write(self.row, DATA_START_COL, LAST_COL, subtitle, font=FONT_BOLD, fill=LABEL_FILL)
                self.row += 1
                if not header_written:
                    self._write_table_header_row(STEP_COLUMNS)
                    header_written = True
                steps = _sort_steps(steps_by_group.get(gid, []))
                if steps:
                    for step in steps:
                        self._write_table_data_row(STEP_COLUMNS, [
                            _text(step.get("stage_name")),
                            _text(step.get("date_actual")),
                            _text(step.get("day_relative")),
                            _resolve_antigen_display(antigen_map, step.get("antigen_id")),
                            _text(step.get("antigen_dose")),
                            _text(step.get("adjuvant_name")),
                            _text(step.get("cpg_dose")),
                            _text(step.get("injection_volume")),
                            _text(step.get("route")),
                            _text(step.get("injection_site")),
                            _text(step.get("remark")),
                        ])
                else:
                    self._write_table_data_row(STEP_COLUMNS, ["-"] * len(STEP_COLUMNS))

        self._write_section_sidebar(section_start, self.row - 1, "免疫方案")

    def _write_section_table(self, section_title: str, columns: list[TableCol], rows: list[list[str]]) -> None:
        section_start = self.row
        self._write_table_header_row(columns)
        if rows:
            for row_values in rows:
                self._write_table_data_row(columns, row_values)
        else:
            self._write_table_data_row(columns, ["-"] * len(columns))
        self._write_section_sidebar(section_start, self.row - 1, section_title)

    def _write_table_header_row(self, columns: list[TableCol]) -> None:
        for start_col, end_col, label in columns:
            self._merge_write(self.row, start_col, end_col, label, font=FONT_BOLD, fill=HEADER_FILL, alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))
        self.ws.row_dimensions[self.row].height = 22
        self.row += 1

    def _write_table_data_row(self, columns: list[TableCol], values: list[str]) -> None:
        for idx, (start_col, end_col, _) in enumerate(columns):
            value = values[idx] if idx < len(values) else "-"
            self._merge_write(
                self.row, start_col, end_col, value,
                alignment=Alignment(
                    horizontal="left" if end_col > start_col else "center",
                    vertical="center",
                    wrap_text=True,
                ),
            )
        self.row += 1

    def _write_section_sidebar(self, start_row: int, end_row: int, title: str) -> None:
        end_row = max(start_row, end_row)
        self.ws.merge_cells(start_row=start_row, start_column=SECTION_COL, end_row=end_row, end_column=SECTION_COL)
        cell = self.ws.cell(row=start_row, column=SECTION_COL, value=title)
        cell.font = FONT_SIDEBAR
        cell.fill = LABEL_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)

    def _write_footer(self) -> None:
        self._merge_write(self.row, SECTION_COL, SECTION_COL + 6, "PM签字 / 日期：", font=FONT_BOLD, height=30)
        self._merge_write(self.row, SECTION_COL + 7, LAST_COL, "方案复核人签字 / 日期：", font=FONT_BOLD)
        self.row += 1

    def _merge_write(
        self,
        row: int,
        col_start: int,
        col_end: int,
        value: Any,
        *,
        font: Font | None = None,
        fill: PatternFill | None = None,
        alignment: Alignment | None = None,
        height: float | None = None,
    ) -> None:
        if col_start != col_end:
            self.ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        cell = self.ws.cell(row=row, column=col_start, value=value)
        cell.font = font or FONT_BODY
        if fill:
            cell.fill = fill
        cell.alignment = alignment or Alignment(horizontal="left", vertical="center", wrap_text=True)
        if height is not None:
            self.ws.row_dimensions[row].height = height


_SCHEME_PDF_NOT_FOUND = "未找到可导出的项目"

# 固定安装路径（Windows 开发机 / Linux bbctg 服务器实测，不用 PATH 猜测）
_SOFFICE_WINDOWS = Path(r"C:\Program Files\LibreOffice\program\soffice.exe")
_SOFFICE_LINUX = Path("/usr/lib/libreoffice/program/soffice")


def export_scheme_pdf_response(db: Session, project_ids: list[int]) -> tuple[BytesIO, str]:
    unique_ids = list(dict.fromkeys(int(pid) for pid in project_ids))
    if len(unique_ids) != 1:
        raise ValueError("打印仅支持单个项目")
    project = load_scheme_data(db, unique_ids[0])
    if not project:
        raise ValueError(_SCHEME_PDF_NOT_FOUND)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    pdf_bytes = _xlsx_to_pdf(_build_project_workbook(project).getvalue())
    return BytesIO(pdf_bytes), _scheme_filename(project, "pdf", timestamp)


def _resolve_soffice() -> Path:
    soffice = _SOFFICE_WINDOWS if sys.platform.startswith("win") else _SOFFICE_LINUX
    if soffice.is_file():
        return soffice
    raise RuntimeError(f"未找到 LibreOffice：{soffice}")


def _xlsx_to_pdf(xlsx_bytes: bytes) -> bytes:
    soffice = _resolve_soffice()
    tmp_base = get_settings().repository_root / "tmp" / "scheme_export"
    tmp_base.mkdir(parents=True, exist_ok=True)
    try:
        with tempfile.TemporaryDirectory(dir=tmp_base) as tmp:
            tmp_path = Path(tmp)
            profile_dir = tmp_path / "lo_profile"
            profile_dir.mkdir()
            xlsx_path = tmp_path / "scheme.xlsx"
            xlsx_path.write_bytes(xlsx_bytes)
            subprocess.run(
                [
                    str(soffice),
                    "--headless",
                    f"-env:UserInstallation={profile_dir.as_uri()}",
                    "--convert-to",
                    "pdf",
                    "--outdir",
                    str(tmp_path),
                    str(xlsx_path),
                ],
                check=True,
                timeout=30,
                capture_output=True,
            )
            return (tmp_path / "scheme.pdf").read_bytes()
    except subprocess.CalledProcessError as exc:
        raise RuntimeError("PDF 转换失败") from exc
    except FileNotFoundError as exc:
        raise RuntimeError("LibreOffice 未生成 PDF") from exc
